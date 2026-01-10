from typing import Iterator
from pathlib import Path


from openpyxl import load_workbook
import json


from ..data.paths import FARMS_SHEET_PATH, GAME_OBJECTS
from ..utils.inputter import inputter
from .actions import Farm
from .objects import GameObject
from . import objects



def farm_generator() -> Iterator[Farm]:
    sheet = load_workbook(FARMS_SHEET_PATH).active
    for row in sheet.iter_rows(min_row=inputter("enter from which google do we start: ", 1) + 1, values_only=True):
        yield Farm(*row)


def load_game_objects():
    with GAME_OBJECTS.open() as f:
        raw_data: dict[str, dict[str, tuple[int, int] | int | str | float]] = json.load(f)

    for object_name, arguments in raw_data.items():
        setattr(
            objects,
            object_name,
            GameObject(
                **{
                    argument_name: argument_val
                    for argument_name, argument_val in arguments.items()
                }
            )
        )
