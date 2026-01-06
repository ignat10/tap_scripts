from typing import Iterator


from openpyxl import load_workbook


from ..data.paths import FARMS_SHEET_PATH
from ..utils.inputter import inputter
from .actions import Farm



sheet = load_workbook(FARMS_SHEET_PATH).active


def farm_generator() -> Iterator[Farm]:
    for row in sheet.iter_rows(min_row=inputter("enter from which google do we start: ", 1) + 1, values_only=True):
        yield Farm(*row)
