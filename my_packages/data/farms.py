from openpyxl import load_workbook


from ..game_tools.actions import Farm
    
from .paths import FARMS_SHEET_PATH


sheet = load_workbook(FARMS_SHEET_PATH).active

max_row = sheet.max_row
max_column = sheet.max_column


from my_packages.utils.inputter import farm_number
farms = [
    Farm(*row)
    for row in sheet.iter_rows(max_row=max_row, max_col=max_column, values_only=True)
    ][farm_number(1)::]
