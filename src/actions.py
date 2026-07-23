from datetime import timedelta
from time import sleep
from typing import Iterator, cast, SupportsInt

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.formula import DataTableFormula, ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet
from keyboard import send
from screen_objects import ScreenObject, reset_screen, back, screenshot, SwipeSpeed, Direction, tap_center

from .objects import objects, resources_technology, equipment, ScreenObjectNames
from .paths import FARMS_SHEET_PATH
from .status import Status, CastleStatus, MapStatus, MineType, check_map_or_castle, check_castle_status, check_map_status
from .utils import object_from_str, log_raise

none_type = type(None)

MAX_MINE_LV = 6
ELITE_MINES = range(10)

def shake() -> None:
    send("f9")

def cell_assert(cell: Cell, typ: type | tuple[type, type]) -> None:
    val = cell.value
    assert isinstance(val,typ), f"cell at {FARMS_SHEET_PATH} {cell.coordinate} should be {typ.__name__}, got '{val.__repr__() if not isinstance(val, (timedelta, DataTableFormula, ArrayFormula, none_type)) else "value doesn't impl repr method"}'"  # type: ignore


class Castle:
    def __init__(self, name: Cell, lv: Cell, google: Cell, account: Cell, alliance: Cell, marches_limit: Cell):
        if lv.value is None:
            lv.value = 1
            save_workbook()

        cell_assert(name, str)
        cell_assert(lv, SupportsInt)
        cell_assert(google, (SupportsInt, none_type))
        cell_assert(account, (SupportsInt, none_type))
        cell_assert(alliance, (str, none_type))
        cell_assert(marches_limit, (int, none_type))

        self.name_cell = name
        self.google_cell = google
        self.account_cell = account
        self.lv_cell = lv
        self.alliance_cell = alliance
        self.max_marches_cell = marches_limit

        self.shook = False
        self.stamina = True
        self.mine_lv = MAX_MINE_LV
        self.mine_type: MineType = MineType.IRON
        self.is_enough_troops = True
        self.elite_mines = self.alliances_elite_mines.setdefault(cast(str, alliance.value), iter(ELITE_MINES))

    alliances_elite_mines: dict[str, Iterator[int]] = {}

    @property
    def name(self) -> ScreenObjectNames:
        val = self.name_cell.value
        return cast(ScreenObjectNames, val)

    @name.setter
    def name(self, value: str):
        self.name_cell.value = value
        save_workbook()

    @property
    def lv(self) -> int:
        return int(cast(SupportsInt, self.lv_cell.value))

    @lv.setter
    def lv(self, value: int):
        self.lv_cell.value = value
        save_workbook()

    @property
    def google(self) -> int:
        return int(cast(SupportsInt, self.google_cell.value))

    @google.setter
    def google(self, value: int):
        self.google_cell.value = value
        save_workbook()

    @property
    def account(self) -> int:
        return int(cast(SupportsInt, self.account_cell.value))

    @account.setter
    def account(self, value: int):
        self.account_cell.value = value
        save_workbook()

    @property
    def alliance(self) -> str:
        return cast(str, self.alliance_cell.value)

    @alliance.setter
    def alliance(self, value: str):
        self.alliance_cell.value = value
        save_workbook()

    @property
    def marches(self) -> int:
        """gets available marches value"""
        cell = self.max_marches_cell
        val = cell.value
        if val is None:
            self.check_marches()
            val = cell.value
        assert isinstance(val, int)
        return val + 1

    @marches.setter
    def marches(self, value: int):
        """sets marches value and saves to .xlsx"""
        self.max_marches_cell.value = value
        save_workbook()

    @staticmethod
    def close_bella():
        while objects['bella'].tap():
            sleep(0.5)

    def new_account(self) -> None:
        """creates new account, upgrades castle to level 4. from city or map."""
        if objects['avatar'].tap():
            objects['account'].waitap()
            objects['new_game'].waitap()
            objects['confirm'].waitap()
            objects['realm'].waitap()
            print("account created")
            sleep(6)

        objects['man'].wait()
        objects['man'].swipe(Direction.Up, SwipeSpeed.Slow, 1.5)
        objects['man'].swipe(Direction.Right, SwipeSpeed.Slow, 0.6)
        sleep(6)
        objects['man'].swipe(Direction.Up, SwipeSpeed.Slow, 0.8)
        objects['man'].swipe(Direction.Right, SwipeSpeed.Slow, 0.6)

        def challenge():
            objects['level'].waitap()
            objects['challenge'].waitap()

        self.kill_monsters()
        print("finished 0 level")
        objects['bella'].wait()
        objects['bella'].spam_tap(4, 0.4)
        challenge()  # first level
        self.kill_monsters()
        print("finished 1st level")
        objects['bella'].waitap()
        objects['bella'].tap()
        challenge()  # second level
        self.kill_monsters()
        print("finished 2nd level")
        objects['bella'].waitap()
        objects['bella'].tap()
        objects['backhand'].waitap()
        objects['first_castle'].waitap()
        objects['upgrade'].waitap()
        objects['upgrade_blue'].waitap()
        self.lv = 2
        objects['bella'].waitap()
        objects['bella'].tap()
        objects['new_monster'].waitap()
        challenge()  # third level
        self.kill_monsters()
        print("finished 3rd level")
        objects['backhand'].waitap()
        objects['bella'].waitap()
        objects['bella'].tap()
        objects['kingroad'].waitap()
        objects['kingroad_go'].waitap()
        objects['hand'].waitap()
        objects['hand'].waitap()
        objects['upgrade_blue'].waitap()
        objects['kingroad'].waitap()
        self.kingroad_claim()
        self.to_map()
        self.to_castle()
        self.upgrade_castle()
        print("upgraded castle to level 3")
        objects['kingroad'].waitap()
        self.kingroad_claim()
        objects['kingroad_go'].waitap()  # barracks task
        objects['bella'].waitap()
        objects['bella'].tap()
        objects['hand'].waitap()
        objects['hand'].waitap()
        objects['kingroad'].waitap()
        self.kingroad_claim()
        objects['kingroad_go'].waitap()  # savior of order task
        objects['hand'].waitap()
        objects['level'].waitap()  # 4th level
        objects['bright_challenge'].waitap()
        self.kill_monsters()
        print("finished 4th level")
        objects['bella'].waitap()
        objects['bella'].tap()
        objects['bella'].tap()
        objects['heroic_evolution'].waitap()
        objects['evolve'].waitap()
        back()
        objects['level'].waitap()
        objects['bright_challenge'].waitap()
        self.kill_monsters()
        print("finished 5th level")
        objects['heroic_evolution'].waitap()
        objects['evolve'].waitap()
        back()
        back()
        self.to_map()
        self.to_castle()
        self.upgrade_castle()
        self.upgrade_castle()
        print("link account!")

    @staticmethod
    def kill_monsters() -> None:
        def bonus_or(direction: Direction):
            if objects['blue_bonus'].tap():
                sleep(0.8)
            if objects['confirm_bonus'].tap():
                sleep(0.8)
            objects['man'].swipe(direction, SwipeSpeed.Slow, 5)
            reset_screen()
        while not (objects['quest_complete'].tap() or objects['quit'].tap()):
            bonus_or(Direction.Right)
            bonus_or(Direction.Up)
            bonus_or(Direction.Left)
            bonus_or(Direction.Down)
            reset_screen()
        sleep(1)

    def check_level(self) -> None:
        objects['avatar'].waitap(2)
        sleep(0.5)
        for lv in range(4, 20):
            if object_from_str(f'castle_level_{lv}').exists():
                self.lv = lv
                back()
                sleep(0.2)
                break
        else:
            log_raise("not recognized any level.")

    def check_marches(self) -> None:
        objects['lord_info'].waitap()
        objects['check_details'].waitap()
        sleep(1)
        for i in reversed(range(4)):
            if object_from_str(f'march_limit_{i}').exists():
                self.marches = i
                save_workbook()
                break
        else:
            log_raise("No march num found in march_limit.")
        back()
        sleep(0.3)
        back()
        sleep(0.2)

    def kingroad_task(self) -> None:
        self.kingroad_claim()
        if objects['kingroad'].tap():
            sleep(1)
        if objects['kingroad_go'].waitap(8):
            sleep(0.5)
            if objects['loading'].exists():
                objects['book'].wait()
            self.close_bella()
            while objects['hand'].waitap(1) or objects['map_hand'].tap():
                print("tapped hand")
                if objects['heroic_evoluation_blue'].waitap(0.7):
                    objects['evolve'].waitap(5)
                objects['go_blue'].tap()
                objects['free'].tap()

            print("no more hands")
            if objects['arrow'].spam_tap(2, 0.5):
                while not objects['attack'].waitap(1.5):
                    tap_center()
                objects['set_out'].waitap(3)

            if objects['check'].exists():
                objects['gather'].waitap(10)
                sleep(1)
                objects['gather'].waitap(4)
                objects['set_out'].waitap(5)
                back()

            if objects['alliance_bonuses'].exists():
                back()
                sleep(1)
            if objects['join'].tap():
                back()
            else:
                objects['apply'].tap_each()

            if objects['bright_challenge'].tap():
                sleep(2)
            if objects['man'].exists():
                self.kill_monsters()

            if objects['research'].exists():
                if not self.speed_up():
                    self.research()

            if objects['forge'].exists():
                self.forge()

            if objects['recruit'].tap() or objects['upgrade'].tap():
                sleep(1)

            objects['free'].tap() or objects['upgrade_blue'].tap() or objects['recruit_blue'].tap()
            objects['get_now'].tap()

            if objects['unlock'].tap():
                print("beast unlocked")
                sleep(10)

            if objects['unlock_land'].tap():
                sleep(0.5)

            self.speed_up()
            self.close_ad()

    @classmethod
    def kingroad_claim(cls):
        """claims completed kingroad tasks"""
        if objects['kingroad'].tap():
            sleep(0.8)
        while objects['kingroad_claim'].tap():
            sleep(1)
            back()
            sleep(0.5)
        if objects['kingroad_done'].tap():
            sleep(2)
            back()
            sleep(0.3)
            cls.close_ad()

    def log_into_account(self) -> None:
        """logs into current account. from city or map."""
        print(f"checking is current castle: {self.name}")
        if not objects[self.name].exists():
            print(f"logging into {self.name}")

            while True:
                if objects["avatar"].tap():
                    sleep(1)
                if objects["account"].tap():
                    sleep(1)
                if objects["switch"].tap():
                    sleep(1)
                objects["login"].waitap(3)
                if not objects['gmail'].wait(10):
                    back()
                    continue
                objects["gmail"].tap_nth(self.google)
                if not objects["acc_list"].wait(15):
                    back()
                    continue
                is_green = objects['green_castle'].exists()
                if not objects["castle"].tap_nth(self.account - is_green):
                    screenshot()
                    raise RuntimeError(
                        f"cannot log into castle \n name: {self.name} \n google: {self.google} \n account: {self.account} \n check screen.png for more info")
                objects["confirm"].waitap()
                break
            print("logged in.")
            sleep(5)

            while check_castle_status() == CastleStatus.NOT_IN_CASTLE:
                reset_screen()
                sleep(1)
                print("loading...")
                sleep(1)
            print("loaded.")
        else:
            print(f"already logged into {self.name}")

    @staticmethod
    def close_ad() -> None:
        """closes ad. from city"""
        if check_castle_status() == CastleStatus.CLOSED_AD:
            return
        reset_screen()
        while check_castle_status() != CastleStatus.CLOSED_AD:
            objects['bella'].spam_tap(5, 0.3)
            objects['continue_game'].tap()
            objects['x'].tap()
            objects['x_new'].tap()
            objects['x_news'].tap()
            objects['claim_daily'].tap()
            if check_castle_status() != CastleStatus.CLOSED_AD:
                for _ in range(3):
                    back()
                    sleep(0.1)
                sleep(0.4)
            reset_screen()
            if check_castle_status() == CastleStatus.CLOSED_AD:
                return
            tap_center()
            objects['map'].wait(1.5)
        sleep(0.5)
        print("ad closed.")

    def claim_rss(self):
        if not self.shook:
            shake()
            sleep(3)
            self.shook = True

    @staticmethod
    def claim_quest():
        if objects['quest'].tap():
            sleep(1)
            while objects['daily_quest_claim'].waitap(0.4):
                back()
            while objects['claim_daily_quest'].waitap(0.4):
                back()
            sleep(1.5)
            objects['growth_quest'].tap()
            for _ in range(3):
                while objects['claim_growth_quest'].waitap(3):
                    if not objects['reward'].waitap(10):
                        back()
                        sleep(1)
                        back()
                if not objects['another_growth_quest'].waitap(2):
                    break
            back()

    def claim(self) -> None:
        """claims recruited troops, gift, and RSS. from city"""
        self.close_ad()
        if objects['horse'].exists():
            print("claiming horses")
            objects['horse'].tap_each()
        if objects['claim'].tap():
            sleep(1)
            back()
            sleep(1)
        objects['help'].tap()
        self.claim_rss()

    @staticmethod
    def events():
        if objects['events'].waitap(1):
            sleep(1)
        for n in range(objects['event'].count()):
            objects['event'].tap_nth(n)
            sleep(1)
            objects['event_claim'].tap_each()
            for _ in range(objects['!'].count()):
                objects['!'].tap()
                sleep(0.5)
                def claims():
                    count = objects['event_claim'].count()
                    objects['event_claim'].tap_each()
                    sleep(1)
                    for _ in range(count):
                        back()
                        sleep(0.3)
                claims()
                claims()
            while objects['event_arrow'].tap():
                sleep(1)
            back()
            sleep(0.5)

    @staticmethod
    def lord_skills() -> None:
        """use lord skills, harvest, gather speed up, recall all. from city or map"""
        print("lord skills...")
        objects["lord"].tap()
        if objects['gather_speed_up'].waitap(2):
            objects['use'].waitap(5)
        objects["harvest"].waitap(3)
        if objects["use"].waitap(2):
            print("harvested")
        objects["recall_all"].waitap(3)
        sleep(0.1)
        if not objects["use"].waitap(2):
            back()
            objects['use'].waitap(2)
        print("lord skills done.")
        reset_screen()
        sleep(0.8)

    @classmethod
    def heal(cls) -> None:
        """heal troops in hospital and sanctuary, then claim healed. from castle."""
        if objects['claim_healed'].tap():
            print("claimed healed")
            sleep(1)
        if objects["hospital"].waitap(0.2):
            print("healing...")
            objects["heal"].waitap()
            if objects['confirm_rss'].waitap(1.5):
                sleep(1)
        if objects["ask_help"].waitap(0.2):
            sleep(1.6)
        if objects['sanctuary'].waitap(0.2):
            print("sanctuary...")
            objects['revive'].waitap()
            objects['claim_holy_water'].waitap(1)
            if not objects['confirm_claim_water'].waitap(0.5):
                objects['holy_quest'].waitap(0.5)
                objects['claim_holy_quest'].waitap(1)
                if objects['confirm_claim_water'].waitap(1):
                    sleep(0.5)
                objects['holy_revival'].waitap(0.8)
            objects['revive'].waitap(1)
            sleep(0.3)
            back()
            sleep(0.8)
        if objects['hospital_building'].waitap(0.2):
            sleep(1)
        cls.speed_up()
        objects['claim_healed'].tap()

    @staticmethod
    def speed_up() -> bool:
        if objects['speed_up'].tap() or objects['speed_up_blue'].tap():
            sleep(0.5)
        if objects['one-tap_speed_up'].tap():
            if objects["confirm_speed_up"].waitap(3):
                sleep(1)
                return True
        return False

    @classmethod
    def to_castle(cls) -> None:
        """goes to castle. from map."""
        print("going inside...")
        while check_map_or_castle() != Status.INSIDE:
            back()
        print("inside")
        cls.close_ad()

    def research(self) -> None:
        marches = self.marches
        if not objects['research'].tap():
            if not objects['college'].tap() and objects['research'].waitap(1):
                self.to_map()
                self.to_castle()
            if objects['college'].tap():
                if not objects['research'].waitap(3):
                    log_raise("not found research button.")
            else:
                log_raise("not found college.")
        done = False
        match marches:
            case 1:
                if self.lv >= 5:
                    print("unlocking 1st additional marche")
                    objects['military'].waitap()
                    if not objects['legion'].waitap(0.5):
                        if not objects['expansion'].tap():
                            if not objects['draft'].tap():
                                log_raise("not found what to research.")
                    done = True
            case 2:
                if self.lv >= 12:
                    print("unlocking 2nd additional marche")
                    objects['military'].waitap()
                    objects['column'].swipe(Direction.Up, SwipeSpeed.Normal, 1)
                    if not objects['legion'].waitap(0.5):
                        if not objects['leadership'].tap():
                            if not objects['horseshoes'].tap():
                                log_raise("not found what to research.")
                    done = True
            case 3:
                if self.lv >= 19:
                    print("unlocking 3rd additional marche")
                    objects['military'].waitap()
                    objects['column'].swipe(Direction.Up, SwipeSpeed.Turbo, 0.7)
                    if not objects['legion'].waitap(0.5):
                        if not objects['horseshoes'].tap():
                            if not objects['expansion'].tap():
                                if not objects['draft'].tap():
                                    log_raise("not found what to research.")
                    done = True
            case n if n != 4:
                log_raise(f"marches should be in range 1..4, got {n}")

        if done:
            objects['research_blue'].waitap(1)
            back()
            back()
            return

        print("researching resources techno.")
        objects['resources'].waitap(10)
        sleep(1)
        for techno in resources_technology:
            techno.waitap(3)
            if objects['research_blue'].waitap(1.5):
                break
            else:
                back()
        back()
        back()

    @staticmethod
    def forge():
        if objects['forge'].tap():
            sleep(0.5)
        max_item: ScreenObject | None = None
        max_num = 0
        for item in equipment:
            item.tap()
            sleep(0.5)
            n = objects['forge_green'].count()
            if n >= max_num:
                max_num = n
                max_item = item
            back()
            sleep(0.5)
        if max_item is not None:
            max_item.tap()
            objects['forge_green'].wait(10)
            objects['forge_green'].tap_nth(max_num - 1)
            if objects['+'].waitap(1):
                objects['select'].waitap(3)
            objects['forge_blue'].waitap(1)
            back()
            back()
            back()

    @classmethod
    def _build_need(cls) -> None:
        """builds required for upgrade buildings. from upgrade menu."""
        if objects['free'].tap() or objects['upgrade_blue'].tap():
            sleep(0.5)
        else:
            if objects['go_upgrade'].tap():
                sleep(0.5)
                cls._build_need()
            else:
                screenshot()
                raise RuntimeError("cannot find upgrade buttons. check screen.png")

    def upgrade_castle(self):
        """upgrades castle or required buildings. from city."""
        objects['castle_building'].waitap()
        objects['upgrade'].waitap()
        sleep(1)
        if self.lv == 2:
            sleep(3)
        if objects['free'].tap() or objects['upgrade_blue'].tap():
            self.lv += 1
        else:
            self._build_need()

    def build(self):
        objects['tasks'].tap()
        sleep(0.8)
        objects['build_task'].tap_nth(0)
        sleep(0.5)
        objects['hand'].spam_tap(2, 0.5)
        objects['upgrade'].tap()
        sleep(0.7)
        self._build_need()

    @staticmethod
    def recruit():
        """recruits horses. from the city."""
        objects['tasks'].tap()
        sleep(0.8)
        for i in range(objects['recruit_task'].count()):
            objects['recruit_task'].tap_nth(i)
            objects['hand'].waitap()
            objects['recruit'].waitap()
            if objects['x_news'].tap():
                sleep(1)
            objects['cavalry'].waitap()
            sleep(0.8)
            objects['previous'].spam_tap(8, 0.02)
            sleep(0.2)
            objects['second'].tap()
            sleep(0.2)
            if objects['recruit_blue'].tap():
                sleep(0.1)
            back()
            objects['tasks'].waitap()
            sleep(1)
        back()
        sleep(0.3)

    @classmethod
    def to_map(cls) -> None:
        """Goes to map from inside city"""
        print("going outside...")
        while not objects['book'].exists():
            if objects["map"].tap() and objects['loading'].wait(1):
                objects['book'].wait()
                break
            else:
                cls.close_ad()
        print("outside.")

    def free_marches(self) -> int:
        """get number of available marches of current castle"""
        limit = self.marches
        if objects['more_marches'].tap():
            sleep(0.3)
        busy = objects['withdraw'].count() + objects['speed_up_march'].count()
        print(f"free marches: {limit - busy}")
        return limit - busy

    def kill_monster(self) -> None:
        if not self.stamina:
            return
        self.to_map()
        objects['search'].tap()
        objects['monster'].waitap(2)
        objects['go'].waitap(2)
        objects['arrow'].wait()
        objects['arrow'].spam_tap(2, 0.2)
        while not objects['attack'].waitap(0.5):
            tap_center()
        objects['set_out'].waitap(2)
        sleep(0.5)
        if check_map_status() == MapStatus.NOT_AT_MAP:
            self.stamina = False
            back()
            back()

    def get_std_mine(self) -> None:
        """Gets standard mine from the map."""

        objects["search"].tap()
        for _ in range(MAX_MINE_LV * MineType.IRON):
            print(f"searching mine. lv {self.mine_lv} {self.mine_type.name.lower()}")
            object_from_str(f"{self.mine_type.name.lower()}_type").waitap()
            objects["plus"].spam_tap(5, 0)
            objects["minus"].spam_tap(6 - self.mine_lv, 0)
            objects["go"].spam_tap(4, 0.1)
            objects['gather'].wait(1.5)

            match check_map_status():
                case MapStatus.FOUND:
                    break
                case MapStatus.NOT_FOUND:
                    if self.mine_type > 1:
                        self.mine_type = MineType(self.mine_type - 1)
                    else:
                        self.mine_lv -= 1
                        self.mine_type = MineType.IRON
                case MapStatus.NOT_AT_MAP:
                    raise RuntimeError(f"Not at map when searching mine.")
        else:
            screenshot()
            raise RuntimeError(f"cannot find standard mine. check screen.png")

        objects['gather'].tap()
        sleep(0.2)
        objects["gather"].tap()
        objects["set_out"].waitap()
        sleep(0.5)
        if check_map_status() == MapStatus.NOT_AT_MAP:
            back()
            print("not enough horses")
            self.is_enough_troops = False
            sleep(1)
        else:
            print("mine taken.")

    def get_elite_mine(self) -> bool:
        """Gets elite mine from the map."""
        print("Elite")
        for e in self.elite_mines:
            assert (
                       status := check_map_or_castle()) == Status.OUTSIDE, f"unexpected status while getting elite mine: {status}"
            objects["book"].tap()
            if objects['x_news'].waitap(0.5):
                sleep(0.5)
            objects["elite_mines"].tap()
            objects['blue'].wait(1)
            if objects["blue"].tap_nth(e):  # color of blue
                if objects["gather"].waitap(2.5):
                    objects["set_out"].waitap()  # regularly I should be there
                    sleep(0.6)
                    if check_map_status() == MapStatus.NOT_AT_MAP:
                        back()
                        sleep(1)
                    return True
            else:
                print("some chemistry error")
                back()
                return False  # if there is no elites
        raise RuntimeError("all elite mines are full.")


workbook = load_workbook(FARMS_SHEET_PATH)
sheet = workbook.active
if sheet is None:
    raise ValueError("No active sheet in workbook")


def save_workbook() -> None:
    workbook.save(FARMS_SHEET_PATH)


def iter_castles():
    assert isinstance(sheet, Worksheet)

    keys: list[str] = [cell.value for cell in sheet[1] if cell.value is not None]  # type: ignore
    inp = input("enter from which castle do we start: ")  # first row is header

    def check_avatar():
        assert isinstance(sheet, Worksheet)

        for cell in sheet['A'][1:]:
            if not isinstance(cell.value, str):
                raise ValueError(f"castle name {cell.value} is not a string")
            if object_from_str(cell.value).exists():
                return cell.row
        else:
            raise RuntimeError("cannot find any castle on the screen")

    match inp:
        case "":
            start_row = check_avatar()
        case "next":
            start_row = check_avatar() + 1
        case n if n.isdigit():
            start_row = max(int(n), 1) + 1  # because of header
        case name:
            start_row = next(
                i for i, cell in enumerate(sheet['A'][1:], start=2)
                if cell.value == name
            )
    for row in sheet.iter_rows(min_row=start_row):
        kwargs = dict(zip(keys, row))
        yield Castle(**kwargs)
